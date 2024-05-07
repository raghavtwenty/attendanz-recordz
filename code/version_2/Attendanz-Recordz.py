"""
All Copyrights Reserved by Attendanz Recordz 2021

Designed and Programmed by Raghava,
Coimbatore Institute Of Technology,
MSc Artificial Intelligence and Machine Learning [Batch - 2021]

Date Created : 20 December 2020 | Last Updated : May 7, 2024
"""

# ----- Importing Required Libraries -----

import os
from pathlib import Path
import tkinter as tk
import webbrowser
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import ttk, scrolledtext, filedialog, messagebox
from datetime import date, time
import time
from fpdf import FPDF


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ----- Create DONT-DELETE.txt File In User Home Directory -----


# ----- Default File Location -----
def default_file_location(Excel_File_Location=" "):
    with open(User_Home_File, "w") as File_Location:
        File_Location.write(Excel_File_Location)


# ----- Initialize Home Directory Path -----
User_Home_Dir = Path.home() / "Attendanz Recordz"
User_Home_File = f"{User_Home_Dir}/Default-File-Location-DONT-DELETE.txt"


# ----- Check For Directory And File -----
if User_Home_Dir.exists() == True:
    if os.path.isfile(User_Home_File) == False:
        default_file_location()
else:
    try:
        # ----- If Not Exist Create New Dir And File -----
        User_Home_Dir.mkdir()
    finally:
        default_file_location()


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ----- Creating Fucntions -----


# ----- About -----
def about_button():

    if about_button:

        # ----- About Window Structure -----
        about_win = tk.Tk()
        about_win.title("About")
        about_win.geometry("500x500")
        about_win.resizable(False, False)
        tk.Label(
            about_win, text="Take online meetings Attendance FASTER THAN EVER!"
        ).pack()

        # ----- About Window Basic Information -----
        tk.Label(about_win, text="App Name: Attendanz Recordz").place(x=0, y=70)
        tk.Label(about_win, text="Last Updated: 27 December 2021").place(x=0, y=100)
        tk.Label(about_win, text="Version: 1.0").place(x=0, y=130)
        tk.Label(
            about_win, text="Compatible with: Mac Os 10.1+ & Windows 7 and above"
        ).place(x=0, y=160)
        tk.Label(about_win, text="© Designed & Programmed by Raghava").place(x=0, y=190)
        tk.Label(about_win, text="Coimbatore Insititute Of Technology,").place(
            x=0, y=210
        )
        tk.Label(
            about_win,
            text="MSc Artificial Intelligence & Machine Learning [Batch - 2021]",
        ).place(x=0, y=230)
        tk.Label(about_win, text="Contact Developer:").place(x=0, y=270)
        tk.Label(about_win, text="@raghavtwenty").place(x=0, y=300)

        # ----- Privacy Policy -----
        tk.Label(
            about_win,
            text="""
\"Transparency Is The Best Privacy Policy\"
We don\'t collect any of our users personal data for analytics,
Neither we track them, This promise will be kept forever.

© All Copyrights Reserved by Attendanz Recordz - 2021
""",
        ).place(x=75, y=380)


# ----- Subject -----
def selecting_subject():
    global subjectchoosen

    subject_var = tk.StringVar()
    tk.Label(root, text="Select Subject :").place(x=0, y=150)
    subjectchoosen = ttk.Combobox(root, width=25, textvariable=subject_var)
    subjectchoosen["values"] = (
        "Computer Science",
        "Office Meetings",
        "Accountancy",
        "Biology",
        "Buisness",
        "Chemistry",
        "Economics",
        "English",
        "Mathmatics",
        "Physics",
        "Science",
        "Social Science",
        "Not Mentioned - Start Typing.",
    )
    subjectchoosen.place(x=120, y=150)
    subjectchoosen.current()


def names_extraction_excel(xlsx_file_name):
    global overall_names_set
    global overall_names_set_len

    wb = load_workbook(xlsx_file_name)
    try:
        extract_names = wb["Sheet1"]
    except:
        messagebox.showwarning("Warning", "Please, Re-name sheet name as 'Sheet1'")
        default_file_location(Excel_File_Location=" ")

    # ----- Extract Data & Convert Into Set -----
    overall_names_set = set([])
    for cell in extract_names["A"]:
        cv = cell.value
        if cv not in [None, " ", "  "]:
            overall_names_set.add(cv.lower())

    overall_names_set_len = len(overall_names_set)

    # ----- Imported File Location -----
    tk.Label(root, text="Imported File Location : ").place(x=0, y=90)
    tk.Label(root, text=":").place(x=0, y=120)
    file_xlsx_location = tk.Entry(root, width=40)
    file_xlsx_location.place(x=7, y=120)
    file_xlsx_location.delete(0, "end")
    file_xlsx_location.insert(0, xlsx_file_name)


# ----- Importing Customized Button -----
def imp_cuz_button():

    # ----- Set Workbook & Worksheet -----
    wb = Workbook()
    ws = wb.active

    if imp_cuz_button:
        # ----- Info Condition -----
        messagebox.showinfo(
            "Import Overall Participants List",
            """Select a Excel sheet (.xlsx) which contains all participants name
sample:
file format: Participants.xlsx
content format:
Participant A (in 1st row)
Participant B (in 2nd row)
...            ...
Note: One participant name in one row and should NOT contain any ','/'.' etc at the end.
""",
        )
        # ----- Opening Xlsx File -----
        file_type_import = [("Excel Spread Sheet", "*.xlsx")]
        file_xlsx_to_set_import = filedialog.askopenfile(
            initialfile="",
            filetypes=file_type_import,
            defaultextension=file_type_import,
        )
        xlsx_file_name = (
            file_xlsx_to_set_import.name
        )  # file.name to avoid iotextwrapper

        default_file_location(
            Excel_File_Location=xlsx_file_name
        )  # Update home file location

        names_extraction_excel(xlsx_file_name)


# process
def process():
    global all_present_set_low
    global all_present_set_low_len
    global absentees_names
    global total_absentees_names_len
    global date_txt
    global time_txt
    global subject_txt
    global grade_txt
    global section_txt
    global total_overall_names_txt
    global total_present_txt
    global total_absentees_txt
    global total_absentees_names_txt
    global present_time

    # present time
    present_time = time.localtime()

    entry_txt_str = str(entry_text_area.get("1.0", "end-1c"))
    all_present_list = entry_txt_str.split("\n")
    while True:
        if get_report:
            break

    # conversion entry text box
    all_present_set_low = set([x.lower() for x in all_present_list])

    all_present_set_low_len = len(all_present_set_low)

    # comparison
    absentees_names_set = overall_names_set.difference(all_present_set_low)
    absentees_names = []
    for ab_name in absentees_names_set:
        if ab_name not in [None, " ", " ", "  "]:
            absentees_names.append(ab_name)

    absentees_names = sorted(absentees_names)
    total_absentees_names_len = len(absentees_names)

    if len(absentees_names) == 0:
        absentees_names = ["NO ABSENTEES."]
        total_absentees_names_len = 0

    # main report
    date_txt = "Date : {}\n".format(date.today())
    report_text_area.insert(tk.END, date_txt)
    report_text_area.see(tk.END)

    time_txt = "Time : {}\n\n".format(time.strftime("%H:%M:%S", present_time))
    report_text_area.insert(tk.END, time_txt)
    report_text_area.see(tk.END)

    subject_txt = "Subject : {}\n".format(subjectchoosen.get())
    report_text_area.insert(tk.END, subject_txt)
    report_text_area.see(tk.END)

    grade_txt = "Grade : {}\n".format(grade_var.get())
    report_text_area.insert(tk.END, grade_txt)
    report_text_area.see(tk.END)

    section_txt = "Section : {}\n\n".format(section_var.get())
    report_text_area.insert(tk.END, section_txt)
    report_text_area.see(tk.END)

    str_overall_names_set_len = str(overall_names_set_len)
    total_overall_names_txt = "Total Number of Participants : {}\n".format(
        str_overall_names_set_len
    )
    report_text_area.insert(tk.END, total_overall_names_txt)
    report_text_area.see(tk.END)

    str_all_present_set_low_len = str(all_present_set_low_len)
    total_present_txt = "Total Number of Participants Present : {}\n".format(
        str_all_present_set_low_len
    )
    report_text_area.insert(tk.END, total_present_txt)
    report_text_area.see(tk.END)

    str_total_absentees_names_len = str(total_absentees_names_len)
    total_absentees_txt = "Total Number of Participants Absent : {}\n\n".format(
        str_total_absentees_names_len
    )
    report_text_area.insert(tk.END, total_absentees_txt)
    report_text_area.see(tk.END)

    report_text_area.insert(tk.END, "Absentees Names :\n")
    report_text_area.see(tk.END)

    for i in absentees_names:
        report_text_area.insert(tk.END, f"{i.upper()}\n")
        report_text_area.see(tk.END)

    # block report box
    report_text_area.configure(state="disabled")


# save pdf
def save_pdf():
    if save_pdf:

        # date_time
        date_pdf = f"Date : {date.today()}"
        time_pdf_strip = time.strftime("%H:%M:%S", present_time)
        time_pdf = f"Time : {time_pdf_strip}"
        dtn = f"Attendanz_Recordz_{date_pdf}_{time_pdf}_{subjectchoosen.get()}.pdf"

        file_type_import = [("Pdf Document", "*.pdf")]
        file_location = filedialog.asksaveasfile(
            initialfile=dtn,
            filetypes=file_type_import,
            defaultextension=file_type_import,
        )
        # pdf file
        file_location_name = file_location.name  # file.name to remove iotextwrapper
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=20)
        pdf.cell(200, 10, txt="Report Generated by Attendanz Recordz", ln=1, align="C")
        pdf.cell(200, 10, txt="", ln=2, align="L")
        pdf.cell(200, 10, txt=date_txt, ln=3, align="L")
        pdf.cell(200, 10, txt=time_txt, ln=4, align="L")
        pdf.cell(200, 10, txt="", ln=5, align="L")
        pdf.cell(200, 10, txt=subject_txt, ln=6, align="L")
        pdf.cell(200, 10, txt=grade_txt, ln=7, align="L")
        pdf.cell(200, 10, txt=section_txt, ln=8, align="L")
        pdf.cell(200, 10, txt="", ln=9, align="L")
        pdf.cell(200, 10, txt=total_overall_names_txt, ln=10, align="L")
        pdf.cell(200, 10, txt=total_present_txt, ln=11, align="L")
        pdf.cell(200, 10, txt=total_absentees_txt, ln=12, align="L")
        pdf.cell(200, 10, txt="", ln=13, align="L")
        pdf.cell(200, 10, txt="Absentees Names : ", ln=14, align="L")

        # itrate absentees names
        for absentees_names_itrate in absentees_names:
            pdf.cell(200, 10, txt=absentees_names_itrate.upper(), ln=15, align="L")

        # ----- PDF Copyright -----
        pdf.cell(200, 10, txt="", ln=15, align="C")
        copyright_txt = [
            "\n",
            "\n",
            "\n",
            "This is a Computer Generated Attendance Report",
            "© Attendanz Recordz, 2021",
            "Designed & Programmed by Raghava",
            "Coimbatore Institute of Technology MSc AI & ML [Batch - 2021]",
            "GitHub: @raghavtwenty",
        ]
        for copyrc in copyright_txt:
            pdf.cell(200, 10, txt=f"{copyrc}\n", ln=10, align="C")

        pdf.output(file_location_name)

        # show final saved location
        report_location = f"Location info : {file_location_name} "
        messagebox.showinfo("Report Saved", report_location)


# get report
def get_report():
    try:
        if get_report:
            global report_text_area

            # result
            tk.Label(root, text="Your Report", font=("Arial", 18, "bold")).place(
                x=675, y=90
            )
            report_text_area = scrolledtext.ScrolledText(
                root, wrap=tk.WORD, width=50, height=25
            )
            report_text_area.place(x=550, y=240)
            report_text_area.focus()

            # processing
            process()

            # save pdf
            tk.Button(root, text="Save Pdf", command=save_pdf).place(x=730, y=645)
    except:
        messagebox.showwarning(
            "Warning", "File Not Selected. Please, Select Overall Participants List."
        )


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ----- Main Window -----
root = tk.Tk()
root.title("Attendanz Recordz")
root.geometry("1080x720")
root.resizable(False, False)
tk.Label(
    root, text="Hello 👋, Welcome To Attendanz Recordz !", font=("Arial", 20)
).pack()

# ----- About Button -----
tk.Button(root, text="About", command=about_button).place(x=1000, y=30)

# ----- Import Customized -----
tk.Button(root, text="Import Overall Participants List", command=imp_cuz_button).place(
    x=0, y=60
)
tk.Button(root, text="Change File", command=imp_cuz_button).place(x=270, y=60)

# ----- Imported File Location -----
tk.Label(root, text="Imported File Location : ").place(x=0, y=90)
tk.Label(root, text=":").place(x=0, y=120)
file_xlsx_location = tk.Entry(root, width=40)
file_xlsx_location.place(x=7, y=120)
file_xlsx_location.delete(0, "end")

# ----- Checking For Default File In Home Directory -----
with open(User_Home_File, "r") as File_Location_Home_Menu:
    default_loc = File_Location_Home_Menu.readline()

if not default_loc == " ":
    try:
        names_extraction_excel(default_loc.strip())
        file_xlsx_location.insert(0, default_loc)
    except FileNotFoundError:
        file_xlsx_location.insert(0, "EXCEL FILE NOT FOUND !!!")
        default_file_location(Excel_File_Location=" ")

else:
    file_xlsx_location.insert(0, "EXCEL FILE NOT FOUND !!!")


# ----- Select Button -----
selecting_subject()

# ----- Select Grade And Section -----
grade_var = tk.StringVar()
tk.Label(root, text="Grade:").place(x=0, y=180)
grade = tk.Entry(root, width=8, textvariable=grade_var).place(x=80, y=180)
section_var = tk.StringVar()
tk.Label(root, text="Section:").place(x=200, y=180)
section = tk.Entry(root, width=8, textvariable=section_var).place(x=290, y=180)

# ----- Participants Names Text Box -----
students_entry = tk.Label(
    root, text="Enter/Paste Current Participants Names* : "
).place(x=0, y=210)
entry_text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=50, height=25)
entry_text_area.place(x=0, y=240)
entry_text_area.focus()

# ----- Get Report -----
tk.Button(root, text="GET REPORT", command=get_report).place(x=180, y=645)

# ----- Me! -----
tk.Label(
    root, text="© Designed & Programmed by Raghava, GitHub: @raghavtwenty"
).place(x=350, y=690)


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# Mainloop

root.mainloop()
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
