# DocString
'''
All Copyrights Reserved by atandanZ recordZ 2021.
Designed and Programmed by raghav, KV Coimbatore - 2020
Github : @raghavtwenty
Created on 14 may 2021
Last updated on 23 may 2021
'''

# -------------------------------
# import
import os
import tkinter as tk
import webbrowser
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import ttk,scrolledtext,filedialog,messagebox
from datetime import date,time
import datetime,time
from fpdf import FPDF

# ----------------------------------------------------------------------------------------------------------------------
# user defined functions

# links
def how_to_use_me():
        webbrowser.open("https://www.google.com")
def website():
        webbrowser.open("https://www.google.com")

# about
def about_button():
        if about_button:

            # about window
            about_win = tk.Tk()
            about_win.title('About')
            about_win.geometry('500x500')
            about_win.resizable(False,False)
            tk.Label(about_win, text = '\"Take online meetings Attandance FASTER THAT EVER!\"').pack()

            # basic information
            tk.Label(about_win, text = 'App Name: atandanZ recordZ').place(x = 0,y = 70)
            tk.Label(about_win, text = 'Last Updated: 21 May 2021').\
                place(x = 0, y = 100)
            tk.Label(about_win, text = 'Version: 1.0').\
                place(x = 0, y = 130)
            tk.Label(about_win, text='User OS: Windows 7 or Later'). \
                place(x=0, y=160)
            tk.Label(about_win, text = 'Designed & Coded by raghav, KV Coimbatore - 2020').place(x = 0, y = 190)
            tk.Label(about_win, text='Looking Tailor Made App For Your School/ Collage/ Organization ?').place(x = 0, y = 250)
            tk.Label(about_win, text='Contact Developer: raghavtwenty@gmail.com').place(x = 0, y = 280)

            # links
            tk.Button(about_win, text  =  'Website', fg = 'blue',
                       command = website).place(x = 0, y = 340)

            # privacy_policy
            tk.Label(about_win, text = '''
\"Transparency Is The Best Privacy Policy\"
We don\'t collect any of our users personal data for analytics,
Neither we track them, This promise will be kept forever.

© All Copyrights Reserved by atandanZ recordZ 2021
''').place(x = 75, y = 380)


# subject
def selecting_subject():
        global subjectchoosen

        subject_var = tk.StringVar()
        subject_label = tk.Label(root, text = 'Select Subject :').place(x = 0, y = 60)
        subjectchoosen = ttk.Combobox(root, width = 46, textvariable = subject_var)
        subjectchoosen['values'] = ('Computer Science', 'Office Meetings','Accountancy',
                                    'Biology', 'Buisness','Chemistry','Economics', 'English',
                                    'Mathmatics', 'Physics','Science', 'Social Science', 'Others')
        subjectchoosen.place(x = 110, y = 60)
        subjectchoosen.current()

# importing customized button
def imp_cuz_button():
        global overall_names_set
        global overall_names_set_len

        # set workbook and worksheet
        wb = Workbook()
        ws = wb.active

        if imp_cuz_button:
            # info condition
            messagebox.showinfo('Import Overall Participants List','''Select a Excel sheet (.xlsx) which contains all participants name
sample:
file format: Participants.xlsx
content format: 
Participant A (in 1st row)
Participant B (in 2nd row)
...            ...
Note: One participant name in one row and should NOT contain any ','/'.' etc at the end.
''')
            # opening xlsx file
            file_type_import = [('Excel Spread Sheet', '*.xlsx')]
            file_xlsx_to_set_import = filedialog.askopenfile(initialfile = '',
                        filetypes = file_type_import,defaultextension = file_type_import)
            xlsx_file_name = file_xlsx_to_set_import.name  # file.name to avoid iotextwrapper
            wb = load_workbook(xlsx_file_name)
            extract_names = wb['Sheet 1']

            # extract data and convert into set
            overall_names_set = set([ ])
            for cell in extract_names['A']:
                overall_names_set.add(cell.value)
            overall_names_set.remove(None)
            overall_names_set_len = len(overall_names_set)

            # file location entry re-configuration
            file_xlsx_location.configure(state = 'normal')
            file_xlsx_location.delete(0, 'end')
            file_xlsx_location.insert(0, xlsx_file_name)
            file_xlsx_location.configure(state = 'disabled')


# process
def process():
        global all_present_set_low
        global all_present_set_low_len
        global absentees_names
        global total_absentees_names_len
        global date_txt
        global time_txt
        global subject_txt
        global grade_txt
        global section_txt
        global total_overall_names_txt
        global total_present_txt
        global total_absentees_txt
        global total_absentees_names_txt

        entry_txt_str = str(entry_text_area.get('1.0', 'end-1c'))
        all_present_list = entry_txt_str.split('\n')
        while True:
            if get_report:
                break

        # conversion entry text box
        all_present_set_low = set([x.lower() for x in all_present_list])
        all_present_set_low_len = len(all_present_set_low)

        # comparison
        absentees_names = overall_names_set.difference(all_present_set_low)
        total_absentees_names_len = len(absentees_names)

        # main report
        date_txt = 'Date:- {}\n'.format(date.today())
        report_text_area.insert(tk.END, date_txt)
        report_text_area.see(tk.END)

        time_txt = 'Time:- {}\n\n'.format(time.strftime("%H::%M::%S", present_time))
        report_text_area.insert(tk.END, time_txt)
        report_text_area.see(tk.END)

        subject_txt = 'Subject:- {}\n'.format(subjectchoosen.get())
        report_text_area.insert(tk.END, subject_txt)
        report_text_area.see(tk.END)

        grade_txt = 'Grade:- {}\n'.format(grade_var.get())
        report_text_area.insert(tk.END, grade_txt)
        report_text_area.see(tk.END)

        section_txt = 'Section:- {}\n\n'.format(section_var.get())
        report_text_area.insert(tk.END, section_txt)
        report_text_area.see(tk.END)

        str_overall_names_set_len = str(overall_names_set_len)
        total_overall_names_txt = 'Total Number of Participants:- {}\n'.format(str_overall_names_set_len)
        report_text_area.insert(tk.END, total_overall_names_txt)
        report_text_area.see(tk.END)

        str_all_present_set_low_len = str(all_present_set_low_len)
        total_present_txt = 'Total Number of Participants Present:- {}\n'.format(str_all_present_set_low_len)
        report_text_area.insert(tk.END, total_present_txt)
        report_text_area.see(tk.END)

        str_total_absentees_names_len = str(total_absentees_names_len)
        total_absentees_txt = 'Total Number of Participants Absent:- {}\n\n'.format(str_total_absentees_names_len)
        report_text_area.insert(tk.END, total_absentees_txt)
        report_text_area.see(tk.END)

        report_text_area.insert(tk.END, 'Absentees Names:-\n')
        report_text_area.see(tk.END)

        absentees_names_str = str(absentees_names)
        total_absentees_names_txt = '{}'.format(absentees_names_str)
        report_text_area.insert(tk.END, total_absentees_names_txt)
        report_text_area.see(tk.END)

        # block report box
        report_text_area.configure(state = 'disabled')


# save pdf
def save_pdf():
        if save_pdf:
            # date_time
            dtn = 'atandanz_Dated_'+str(date.today()) + '_' + str(time.strftime("%H_%M_%S", present_time))

            file_type_import = [('Pdf Document', '*.pdf')]
            file_location = filedialog.asksaveasfile(initialfile=dtn, filetypes=file_type_import,
                                                     defaultextension=file_type_import)
            # pdf file
            file_location_name = file_location.name # file.name to remove iotextwrapper
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size = 20)
            pdf.cell(200, 10, txt = 'Report Generated by atandanZ recordZ', ln=1, align='C')
            pdf.cell(200, 10, txt = '', ln = 2, align = 'L')
            pdf.cell(200, 10, txt = date_txt, ln = 3, align = 'L')
            pdf.cell(200, 10, txt = time_txt, ln = 4, align = 'L')
            pdf.cell(200, 10, txt = '', ln = 5, align = 'L')
            pdf.cell(200, 10, txt = subject_txt, ln = 6, align = 'L')
            pdf.cell(200, 10, txt = grade_txt, ln = 7, align = 'L')
            pdf.cell(200, 10, txt = section_txt, ln=8, align = 'L')
            pdf.cell(200, 10, txt = '', ln = 9, align = 'L')
            pdf.cell(200, 10, txt = total_overall_names_txt, ln = 10, align = 'L')
            pdf.cell(200, 10, txt = total_present_txt, ln = 11, align = 'L')
            pdf.cell(200, 10, txt = total_absentees_txt, ln = 12, align = 'L')
            pdf.cell(200, 10, txt = '', ln = 13, align = 'L')
            pdf.cell(200, 10, txt = 'Absentees Names:-', ln = 14, align = 'L')

            # itrate absentees names
            for absentees_names_itrate in absentees_names:
                pdf.cell(200, 10, txt = absentees_names_itrate, ln = 15, align = 'L')
            pdf.output(file_location_name)

            # show final saved location
            report_location = 'Location info: ' + file_location_name
            messagebox.showinfo('Report Saved',report_location)



# get report
def get_report():
     try:
        if get_report:
            global report_text_area
            global present_time
            global date_today
            global current_time

            # report date
            date_today = 'Date:- '+ str(date.today())
            report_date = tk.Entry(root, width = 20)
            report_date.delete(0, 'end')
            report_date.insert(0, date_today)
            report_date.configure(state = 'disabled')
            report_date.place(x = 550, y = 170)

            # report time
            present_time = time.localtime()
            current_time = 'Time:- '+ time.strftime("%H::%M::%S", present_time)
            report_time = tk.Entry(root, width = 20)
            report_time.delete(0, 'end')
            report_time.insert(0, current_time)
            report_time.configure(state = 'disabled')
            report_time.place(x = 828, y = 170)

            # result
            tk.Label(root, text = 'Your Report',font = ('Arial',18,'bold')).place(x = 675, y = 90)
            report_text_area = scrolledtext.ScrolledText(root, wrap = tk.WORD, width = 50, height = 25)
            report_text_area.place(x = 550, y = 230)
            report_text_area.focus()

            # processing
            process()

            # save pdf
            tk.Button(root, text = 'Save Pdf', command = save_pdf).place(x = 730, y = 645)
     except:
        messagebox.showwarning('Warning','File Not Selected. Please, Select Overall Participants List.')



#---------------main window---------------
root = tk.Tk()
root.title('atandanZ recordZ')
root.iconbitmap(r'C:\Users\91904\AppData\Local\Programs\Python\Python39-32\Scripts\atandanZ_recordZ_icon.ico')
root.geometry('1080x720')
root.resizable(False,False)
tk.Label(root, text = 'Hello, Welcome To atandanZ recordZ !', font = ('Arial',20)).pack()

#how to use
tk.Button(root, text = 'How To Use Me!', fg = 'blue', command = how_to_use_me).\
    place(x = 860,y = 30)

#about button
tk.Button(root, text = 'About', command = about_button).place(x = 1000,y = 30)

#select subject
selecting_subject()

#enter grade and section
grade_var=tk.StringVar()
tk.Label(root, text = 'Grade:').place(x = 0, y = 90)
grade=tk.Entry(root, width = 10, textvariable = grade_var).place(x = 112, y = 90)
section_var=tk.StringVar()
tk.Label(root, text = 'Section:').place(x = 250, y = 90)
section=tk.Entry(root, width = 10, textvariable = section_var).place(x = 342, y = 90)

#import customized
tk.Button(root, text = 'Import Overall Participants List', command = imp_cuz_button).\
    place(x = 0,y = 120)
tk.Button(root, text = 'Change File', command = imp_cuz_button).place(x = 333,y = 120)

#imported location
tk.Label(root, text = 'Imported File Location:').place(x = 0, y = 150)
tk.Label(root, text = ':').place(x = 0, y = 170)
file_xlsx_location = tk.Entry(root, width = 66)
file_xlsx_location.configure(state = 'disabled')
file_xlsx_location.place(x = 7, y = 170)

#enter present names text box
students_entry = tk.Label(root,text = 'Enter/ Paste Today\'s Participants Name* :').place(x = 0, y = 200)
entry_text_area = scrolledtext.ScrolledText(root, wrap = tk.WORD, width = 50, height = 25)
entry_text_area.place(x = 0,y = 230)
entry_text_area.focus()

#get report
tk.Button(root,text='Get report', command=get_report).place(x = 180, y = 645)

#me!
tk.Label(root, text = 'Designed & Coded by raghav, KV Coimbatore - 2020').place(x = 350, y = 690)


#-------------run--------
root.mainloop()
